
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

OVERTIME_YES = "Yes"

def apply_overtime_highlighting(ws):
    header_row = [cell.value for cell in ws[1]]
    try:
        overtime_idx = header_row.index("Overtime")
    except ValueError:
        return
    overtime_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[overtime_idx].value == OVERTIME_YES:
            for cell in row:
                cell.fill = overtime_fill

def export_to_excel(df_jobs, calendar_df, metrics_df, filename="mowing_team_schedule.xlsx"):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_jobs.to_excel(writer, sheet_name='Detailed Assignments', index=False)
        calendar_df.to_excel(writer, sheet_name='Calendar View', index=False)
        metrics_df.to_excel(writer, sheet_name='Metrics Summary', index=False)
    wb = load_workbook(filename)
    apply_overtime_highlighting(wb["Detailed Assignments"])
    wb.save(filename)
