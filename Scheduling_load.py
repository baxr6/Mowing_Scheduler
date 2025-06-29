import pandas as pd
import heapq
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import Patch
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import logging
import sys
import os
from datetime import date, timedelta
import json
from pathlib import Path

# === Logger Setup ===
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s: %(message)s")

# === Load Config ===
def load_config(config_path="config.json"):
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: {config_path}")
    with open(config_path) as f:
        config = json.load(f)
    config["START_DATE"] = date.fromisoformat(config["START_DATE"])
    config["PUBLIC_HOLIDAYS"] = {date.fromisoformat(d) for d in config["PUBLIC_HOLIDAYS"]}
    return config


CONFIG = load_config("/home/deano/Projects/Scheduling/config.json")


DEFAULT_MOWING_RATE_SQM_PER_HOUR = CONFIG["DEFAULT_MOWING_RATE_SQM_PER_HOUR"]
DEFAULT_WORKDAY_HOURS = CONFIG["DEFAULT_WORKDAY_HOURS"]
DEFAULT_WORKDAYS_PER_WEEK = CONFIG["DEFAULT_WORKDAYS_PER_WEEK"]
DEFAULT_BUFFER = CONFIG["DEFAULT_BUFFER"]
START_DATE = CONFIG["START_DATE"]
PUBLIC_HOLIDAYS = CONFIG["PUBLIC_HOLIDAYS"]
TEAM_NAME_MAPPING = CONFIG["TEAM_NAME_MAPPING"]
COMBINED_TEAMS = list(TEAM_NAME_MAPPING.keys())
SUBURB_TO_COMBINED_TEAM = CONFIG["SUBURB_TO_COMBINED_TEAM"]
WEEK_RANGE_TO_INCLUDE = CONFIG.get("WEEK_RANGE_TO_INCLUDE")  # Optional

OVERTIME_YES = "Yes"
OVERTIME_NO = "No"

def get_nth_working_day(start_date, n, skipped_dates):
    current = start_date
    count = 0
    while count < n:
        current += timedelta(days=1)
        if current not in skipped_dates:
            count += 1
    return current

def load_parks_from_csv(csv_path):
    path = Path(csv_path)
    if not path.exists():
        logging.error(f"CSV file not found: {csv_path}")
        sys.exit(1)
    df = pd.read_csv(path)
    required_cols = {'name', 'area_sqm', 'suburb'}
    if not required_cols.issubset(df.columns):
        logging.error(f"CSV missing required columns: {required_cols}")
        sys.exit(1)
    return df.to_dict(orient='records')

def build_team_heap(team_status, team_list):
    heap = []
    for team in team_list:
        status = team_status[team]
        heap.append((status["total_hours_assigned"], len(status["parks_assigned_set"]), team))
    heapq.heapify(heap)
    return heap

def assign_parks_with_combined_teams(parks, combined_teams, team_map, suburb_map,
                                     rate_per_hour, workday_hours, buffer=DEFAULT_BUFFER,
                                     start_date=START_DATE, skipped_dates=PUBLIC_HOLIDAYS):
    parks_sorted = sorted(parks, key=lambda x: x["area_sqm"], reverse=True)
    all_teams = sorted({team for teams in team_map.values() for team in teams})
    team_status = {
        team: {
            "team_name": team,
            "current_day": 1,
            "hours_assigned_today": 0,
            "total_hours_assigned": 0,
            "parks_assigned_set": set(),
            "total_overtime_hours": 0,
            "overtime_jobs": 0
        } for team in all_teams
    }

    jobs = {team: [] for team in all_teams}

    for park in parks_sorted:
        if park["area_sqm"] <= 0:
            continue
        total_time = (park["area_sqm"] / rate_per_hour) * (1 + buffer)
        time_remaining = total_time

        allowed_combined = suburb_map.get(park["suburb"], combined_teams)
        allowed_individual = [t for name in allowed_combined for t in team_map[name]]

        while time_remaining > 0:
            team_heap = build_team_heap(team_status, allowed_individual)
            if not team_heap:
                logging.error(f"No teams available for park: {park['name']}")
                break

            assigned = False
            for _ in range(len(team_heap)):
                _, _, t = heapq.heappop(team_heap)
                status = team_status[t]

                while True:
                    work_date = get_nth_working_day(start_date, status["current_day"], skipped_dates)
                    if work_date not in skipped_dates:
                        break
                    status["current_day"] += 1

                if status["hours_assigned_today"] >= workday_hours:
                    status["current_day"] += 1
                    status["hours_assigned_today"] = 0
                    continue

                available = workday_hours - status["hours_assigned_today"]
                if available <= 0:
                    continue

                time_to_assign = min(available, time_remaining)
                area_chunk = time_to_assign * rate_per_hour / (1 + buffer)
                jobs[t].append({
                    "name": park["name"],
                    "suburb": park["suburb"],
                    "area_sqm": round(area_chunk, 2),
                    "estimated_hours": round(time_to_assign, 2),
                    "day": status["current_day"],
                    "date": work_date.isoformat(),
                    "overtime": OVERTIME_NO
                })
                status["hours_assigned_today"] += time_to_assign
                status["total_hours_assigned"] += time_to_assign
                status["parks_assigned_set"].add(park["name"])
                time_remaining -= time_to_assign
                assigned = True
                break

            if not assigned:
                fallback_heap = build_team_heap(team_status, allowed_individual)
                if not fallback_heap:
                    logging.error(f"No fallback teams for park: {park['name']}")
                    break
                _, _, t = heapq.heappop(fallback_heap)
                status = team_status[t]
                status["current_day"] += 1
                status["hours_assigned_today"] = 0
                work_date = get_nth_working_day(start_date, status["current_day"], skipped_dates)

                jobs[t].append({
                    "name": park["name"],
                    "suburb": park["suburb"],
                    "area_sqm": round(time_remaining * rate_per_hour / (1 + buffer), 2),
                    "estimated_hours": round(time_remaining, 2),
                    "day": status["current_day"],
                    "date": work_date.isoformat(),
                    "overtime": OVERTIME_YES
                })

                status["hours_assigned_today"] += time_remaining
                status["total_hours_assigned"] += time_remaining
                status["parks_assigned_set"].add(park["name"])
                status["total_overtime_hours"] += time_remaining
                status["overtime_jobs"] += 1
                time_remaining = 0

    return jobs

def export_jobs_to_df(jobs):
    rows = []
    for team, assignments in jobs.items():
        for job in assignments:
            rows.append({
                "Team": team,
                "Day": job["day"],
                "Date": job["date"],
                "Park": job["name"],
                "Suburb": job["suburb"],
                "Area (sqm)": job["area_sqm"],
                "Estimated Hours": job["estimated_hours"],
                "Overtime": job["overtime"]
            })
    df = pd.DataFrame(rows)
    df.sort_values(by=["Team", "Day"], inplace=True)
    return df

def add_week_and_weekday(df):
    df = df.copy()
    df['Week'] = ((df['Day'] - 1) // DEFAULT_WORKDAYS_PER_WEEK) + 1
    df['Weekday'] = pd.to_datetime(df['Date']).dt.strftime('%a')
    return df

def build_calendar(df, week_range=None):
    calendar = {}
    for _, row in df.iterrows():
        week = row['Week']
        if week_range and week not in week_range:
            continue
        team = row['Team']
        day = row['Weekday']
        text = f"{row['Park']} ({row['Estimated Hours']}h)"
        calendar.setdefault(team, {}).setdefault(week, {}).setdefault(day, []).append(text)

    all_weeks = sorted({w for t in calendar.values() for w in t})
    weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    rows = []

    for team in sorted(calendar):
        row = {"Team": team}
        for week in all_weeks:
            for day in weekdays:
                key = f"W{week} {day}"
                items = calendar.get(team, {}).get(week, {}).get(day, [])
                row[key] = "\n".join(sorted(items))
        rows.append(row)
    return pd.DataFrame(rows)

def generate_metrics(df):
    overtime = df["Overtime"] == OVERTIME_YES
    summary = df.groupby('Team').agg(
        Total_Parks=('Park', 'nunique'),
        Total_Area_Sqm=('Area (sqm)', 'sum'),
        Total_Hours=('Estimated Hours', 'sum'),
        Days_Worked=('Day', 'nunique'),
        Total_Overtime_Hours=('Estimated Hours', lambda x: x[overtime.loc[x.index]].sum())
    ).reset_index()
    summary["Avg_Hours_Per_Day"] = (summary["Total_Hours"] / summary["Days_Worked"]).round(2)
    return summary.round(2)

def export_to_excel(df_jobs, calendar_df, metrics_df, filename="mowing_team_schedule.xlsx"):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_jobs.to_excel(writer, sheet_name='Detailed Assignments', index=False)
        calendar_df.to_excel(writer, sheet_name='Calendar View', index=False)
        metrics_df.to_excel(writer, sheet_name='Metrics Summary', index=False)

    wb = load_workbook(filename)
    overtime_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws = wb["Detailed Assignments"]
    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.font = Font(bold=True)

    overtime_col = [i for i, cell in enumerate(ws[1], start=1) if cell.value == "Overtime"]
    if overtime_col:
        overtime_idx = overtime_col[0]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[overtime_idx - 1].value == OVERTIME_YES:
                for cell in row:
                    cell.fill = overtime_fill

    wb.save(filename)

def export_gantt_chart(df_jobs, filename="mowing_team_schedule.xlsx"):
    df = df_jobs.copy()
    df["Start"] = pd.to_datetime(df["Date"])
    df["End"] = df["Start"] + pd.to_timedelta(df["Estimated Hours"], unit="h")
    teams = sorted(df["Team"].unique())
    idx = {team: i for i, team in enumerate(teams)}
    fig, ax = plt.subplots(figsize=(15, 0.5 * len(teams) + 3))
    colors = {OVERTIME_YES: "red", OVERTIME_NO: "green"}

    for _, row in df.iterrows():
        ax.barh(
            y=idx[row["Team"]],
            left=row["Start"],
            width=row["End"] - row["Start"],
            height=0.4,
            color=colors.get(row["Overtime"], "blue"),
            edgecolor='black'
        )

    ax.set_yticks(list(idx.values()))
    ax.set_yticklabels(teams)
    ax.xaxis.set_major_locator(mdates.DayLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
    plt.xticks(rotation=45)
    plt.title("Mowing Schedule Gantt Chart")
    plt.tight_layout()
    img_file = "gantt_temp.png"
    plt.savefig(img_file)
    plt.close()

    wb = load_workbook(filename)
    ws = wb.create_sheet("Gantt Chart")
    img = XLImage(img_file)
    img.anchor = 'A1'
    ws.add_image(img)
    wb.save(filename)
    os.remove(img_file)

def main(csv_path="/home/deano/Projects/Scheduling/sample_parks_300.csv"):
    parks = load_parks_from_csv(csv_path)

    jobs = assign_parks_with_combined_teams(
        parks, COMBINED_TEAMS, TEAM_NAME_MAPPING, SUBURB_TO_COMBINED_TEAM,
        DEFAULT_MOWING_RATE_SQM_PER_HOUR, DEFAULT_WORKDAY_HOURS
    )

    df_jobs = export_jobs_to_df(jobs)
    df_jobs = add_week_and_weekday(df_jobs)

    # Week filtering via config
    if WEEK_RANGE_TO_INCLUDE:
        df_jobs_filtered = df_jobs[df_jobs["Week"].isin(WEEK_RANGE_TO_INCLUDE)]
    else:
        df_jobs_filtered = df_jobs

    calendar_df = build_calendar(df_jobs_filtered, week_range=WEEK_RANGE_TO_INCLUDE)
    metrics_df = generate_metrics(df_jobs_filtered)

    export_to_excel(df_jobs_filtered, calendar_df, metrics_df)
    export_gantt_chart(df_jobs_filtered)

    logging.info("\n📅 Calendar View:\n" + calendar_df.head().to_string(index=False))
    logging.info("\n📊 Metrics:\n" + metrics_df.to_string(index=False))

if __name__ == "__main__":
    main()