import pandas as pd
from collections import defaultdict
import heapq
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import logging
import sys
import os

# === Configuration Constants ===
DEFAULT_MOWING_RATE_SQM_PER_HOUR = 1200
DEFAULT_WORKDAY_HOURS = 8
DEFAULT_WORKDAYS_PER_WEEK = 6
DEFAULT_BUFFER = 0.5  # 50% buffer time

TEAM_NAMES = [
    "Southern", "Western", "South Western", "Northern Team",
    "Eastern 1", "Eastern 2", "Eastern 3", "Eastern 4", "Eastern 5"
]

SUBURB_TEAMS_MAP = {
    "Riverview": ["Southern", "South Western"],
    "Flindersview": ["Western", "South Western"],
    "Dinmore": ["Northern Team"],
    "Raceview": ["Eastern 3", "Eastern 4", "Eastern 5"],
    "Springfield": ["Eastern 1", "Eastern 2"],
}

# === Logger Setup ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s: %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)


def load_parks_from_csv(csv_path):
    if not os.path.exists(csv_path):
        logging.error(f"CSV file not found: {csv_path}")
        sys.exit(1)
    df = pd.read_csv(csv_path)
    required_cols = {'name', 'area_sqm', 'suburb'}
    if not required_cols.issubset(df.columns):
        logging.error(f"CSV missing required columns: {required_cols}")
        sys.exit(1)
    return df.to_dict(orient='records')


def build_priority_heap(team_status, allowed_teams):
    """
    Build a priority heap of teams eligible to take the next job chunk.
    Sort by (total_hours_assigned, number of unique parks assigned).
    """
    heap = []
    for t_name in allowed_teams:
        status = team_status[t_name]
        heap.append((status["total_hours_assigned"], len(status["parks_assigned_set"]), t_name))
    heapq.heapify(heap)
    return heap


def assign_parks_with_suburb_constraints(
    parks, team_names, rate_per_hour, workday_hours,
    suburb_teams_map, buffer=DEFAULT_BUFFER
):
    parks_sorted = sorted(parks, key=lambda x: x["area_sqm"], reverse=True)
    
    team_status = {
        name: {
            "team_name": name,
            "current_day": 1,
            "hours_assigned_today": 0,
            "total_hours_assigned": 0,
            "parks_assigned_set": set(),
            "total_overtime_hours": 0,
            "overtime_jobs": 0
        }
        for name in team_names
    }

    jobs = defaultdict(list)

    for park in parks_sorted:
        if park["area_sqm"] <= 0:
            logging.warning(f"Skipping park with non-positive area: {park['name']}")
            continue

        total_time = (park["area_sqm"] / rate_per_hour) * (1 + buffer)
        time_remaining = total_time
        logging.info(f"Assigning park '{park['name']}' ({park['area_sqm']} sqm, est {total_time:.2f}h)")

        allowed_teams = suburb_teams_map.get(park.get("suburb"), team_names)

        if not allowed_teams:
            logging.warning(f"No teams mapped for suburb '{park.get('suburb')}', assigning to all teams.")
            allowed_teams = team_names

        while time_remaining > 0:
            team_heap = build_priority_heap(team_status, allowed_teams)
            assigned_this_round = False

            for _ in range(len(team_heap)):
                _, _, t_name = heapq.heappop(team_heap)
                status = team_status[t_name]

                if status["hours_assigned_today"] >= workday_hours:
                    status["current_day"] += 1
                    status["hours_assigned_today"] = 0

                available_today = workday_hours - status["hours_assigned_today"]
                if available_today == 0:
                    continue

                time_to_assign = min(time_remaining, available_today)
                is_overtime = "No"
                area_chunk = time_to_assign * rate_per_hour / (1 + buffer)

                jobs[t_name].append({
                    "name": park["name"],
                    "suburb": park.get("suburb", ""),
                    "area_sqm": round(area_chunk, 2),
                    "estimated_hours": round(time_to_assign, 2),
                    "day": status["current_day"],
                    "overtime": is_overtime
                })

                status["hours_assigned_today"] += time_to_assign
                status["total_hours_assigned"] += time_to_assign
                status["parks_assigned_set"].add(park["name"])
                time_remaining -= time_to_assign
                assigned_this_round = True

                if time_remaining <= 0:
                    break

            if not assigned_this_round:
                # All teams booked today; assign as overtime to lowest-loaded team
                team_heap = build_priority_heap(team_status, allowed_teams)
                _, _, t_name = heapq.heappop(team_heap)
                status = team_status[t_name]

                status["current_day"] += 1
                status["hours_assigned_today"] = 0

                time_to_assign = time_remaining
                area_chunk = time_to_assign * rate_per_hour / (1 + buffer)

                jobs[t_name].append({
                    "name": park["name"],
                    "suburb": park.get("suburb", ""),
                    "area_sqm": round(area_chunk, 2),
                    "estimated_hours": round(time_to_assign, 2),
                    "day": status["current_day"],
                    "overtime": "Yes"
                })

                status["hours_assigned_today"] += time_to_assign
                status["total_hours_assigned"] += time_to_assign
                status["parks_assigned_set"].add(park["name"])
                status["total_overtime_hours"] += time_to_assign
                status["overtime_jobs"] += 1
                time_remaining = 0

    # Final summary
    logging.info("\n=== TEAM SUMMARY ===")
    for t in team_names:
        status = team_status[t]
        logging.info(
            f"{t}: Days Worked={status['current_day']}, "
            f"Total Hours={status['total_hours_assigned']:.2f}, "
            f"Overtime Hours={status['total_overtime_hours']:.2f}, "
            f"Overtime Jobs={status['overtime_jobs']}, "
            f"Unique Parks={len(status['parks_assigned_set'])}"
        )

    return jobs



def export_jobs_to_df(jobs):
    records = []
    for team, assignments in jobs.items():
        for job in assignments:
            records.append({
                "Team": team,
                "Day": job["day"],
                "Park": job["name"],
                "Suburb": job["suburb"],
                "Area (sqm)": job["area_sqm"],
                "Estimated Hours": job["estimated_hours"],
                "Overtime": job["overtime"]
            })
    return pd.DataFrame(records)


def add_week_and_weekday(df, workdays_per_week=DEFAULT_WORKDAYS_PER_WEEK):
    df = df.copy()
    df['Week'] = ((df['Day'] - 1) // workdays_per_week) + 1
    df['Weekday_Num'] = ((df['Day'] - 1) % workdays_per_week) + 1
    weekday_map = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu", 5: "Fri", 6: "Sat"}
    df['Weekday'] = df['Weekday_Num'].map(weekday_map)
    return df


def build_calendar(df_jobs, workdays_per_week=DEFAULT_WORKDAYS_PER_WEEK):
    calendar = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for _, row in df_jobs.iterrows():
        calendar[row["Team"]][row["Week"]][row["Weekday"]].append(
            f"{row['Park']} ({row['Estimated Hours']}h)"
        )

    all_weeks = sorted({week for team in calendar.values() for week in team})
    weekday_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"][:workdays_per_week]

    rows = []
    for team in sorted(calendar.keys()):
        row = {"Team": team}
        for week in all_weeks:
            for day in weekday_names:
                key = f"W{week} {day}"
                items = calendar[team].get(week, {}).get(day, [])
                row[key] = "\n".join(sorted(items))
        rows.append(row)
    return pd.DataFrame(rows)


def generate_metrics(df_jobs):
    overtime_filter = df_jobs["Overtime"] == "Yes"
    summary = df_jobs.groupby('Team').agg(
        Total_Parks=('Park', 'nunique'),
        Total_Area_Sqm=('Area (sqm)', 'sum'),
        Total_Hours=('Estimated Hours', 'sum'),
        Days_Worked=('Day', 'nunique'),
        Total_Overtime_Hours=('Estimated Hours', lambda x: x[overtime_filter.loc[x.index]].sum())
    ).reset_index()

    summary['Avg_Hours_Per_Day'] = (summary['Total_Hours'] / summary['Days_Worked']).round(2)
    summary['Total_Area_Sqm'] = summary['Total_Area_Sqm'].round(2)
    summary['Total_Hours'] = summary['Total_Hours'].round(2)
    summary['Total_Overtime_Hours'] = summary['Total_Overtime_Hours'].round(2)
    return summary


def export_to_excel(df_jobs, calendar_df, metrics_df, filename="mowing_team_schedule.xlsx"):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_jobs.to_excel(writer, sheet_name='Detailed Assignments', index=False)
        calendar_df.to_excel(writer, sheet_name='Calendar View', index=False)
        metrics_df.to_excel(writer, sheet_name='Metrics Summary', index=False)

    wb = load_workbook(filename)
    overtime_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        if sheet_name == "Detailed Assignments":
            overtime_col_idx = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == "Overtime":
                    overtime_col_idx = idx
                    break
            if overtime_col_idx:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if row[overtime_col_idx - 1].value == "Yes":
                        for cell in row:
                            cell.fill = overtime_fill

    wb.save(filename)
    logging.info(f"📁 Excel file exported: {filename}")


def main():
    parks = [
        {"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"},
	{"name": "Emerald Park", "area_sqm": 2608.16, "suburb": "Flindersview"},
        {"name": "ENCHANTED WOODS PARK", "area_sqm": 8081.72, "suburb": "Springfield"},
        {"name": "Eric Edwardson Park", "area_sqm": 12501.26, "suburb": "Dinmore"},
        {"name": "Eric Street Reserve (a)", "area_sqm": 6870.89, "suburb": "Raceview"}
    ]

    jobs = assign_parks_with_suburb_constraints(
        parks, TEAM_NAMES, DEFAULT_MOWING_RATE_SQM_PER_HOUR,
        DEFAULT_WORKDAY_HOURS, SUBURB_TEAMS_MAP, buffer=DEFAULT_BUFFER
    )

    df_jobs = export_jobs_to_df(jobs)
    df_jobs = add_week_and_weekday(df_jobs, DEFAULT_WORKDAYS_PER_WEEK)
    calendar_df = build_calendar(df_jobs, DEFAULT_WORKDAYS_PER_WEEK)
    metrics_df = generate_metrics(df_jobs)

    logging.info("\n📅 Schedule Preview:\n" + calendar_df.head(10).to_string(index=False))
    logging.info("\n📈 Team Metrics:\n" + metrics_df.to_string(index=False))

    #export_to_excel(df_jobs, calendar_df, metrics_df, "mowing_team_schedule.xlsx")


if __name__ == "__main__":
    main()
